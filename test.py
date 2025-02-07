import os
import pandas as pd
from openpyxl import load_workbook

path = r"C:\Users\admin\Desktop\销售出库单-XSCKD2025020505-314.xlsx"

wb = load_workbook(path)

ws = wb.active

a = ws['I10'].value

print(a)
