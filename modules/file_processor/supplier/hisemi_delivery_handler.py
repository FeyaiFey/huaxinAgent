"""
池州华宇供应商Excel处理器
处理池州华宇供应商的送货单Excel文件
"""

import os
from typing import Dict, List, Optional, Any
from pathlib import Path
from openpyxl import load_workbook
from .base_delivery_handler import BaseDeliveryExcelHandler

class HisemiHandler(BaseDeliveryExcelHandler):
    """
    池州华宇供应商Excel处理器
    
    处理说明:
    1. 固定读取'Page 1'工作表
    2. 日期位置固定在P4单元格
    3. 从第7行开始读取数据，直到遇到'TOTAL'行
    4. 跳过空行（以订单号是否存在为判断依据）
    5. 验证和格式化数据
    6. 将数据保存到JSON文件
    7. 将Excel文件移动到归档目录
    """
    
    def process(self, match_result: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """
        处理池州华宇的Excel文件
        
        Args:
            match_result: 规则引擎匹配结果
            
        Returns:
            Dict: 处理结果
        """
        try:
            # 用于存储所有数据
            all_data: Dict[str, List[Dict[str, Any]]] = {}

            # 确保目录存在
            attachment_folder = match_result.get('actions').get('attachment_folder')
            if not attachment_folder:
                self.logger.error(f"邮件数据中缺少附件目录")
                return None
                
            if not os.path.exists(attachment_folder):
                try:
                    os.makedirs(attachment_folder)
                    self.logger.info(f"已创建目录: {attachment_folder}")
                except Exception as e:
                    self.logger.error(f"创建目录失败: {attachment_folder} - {str(e)}")
                    return None
                
            # 遍历目录下的所有文件
            excel_count = 0
            processed_count = 0

            for file in os.listdir(attachment_folder):
                if not file.lower().endswith(('.xlsx', '.xls')):
                    continue
                    
                file_path = os.path.join(attachment_folder, file)
                try:
                    self.logger.info(f"开始处理池州华宇送货单: {Path(file_path).name}")
                    data_dict = self._process_huayu_return_dict(file_path)
                    excel_count += 1
                    
                    if not data_dict:
                        self.logger.warning(f"文件处理未返回数据: {file}")
                        continue
                        
                    # 验证和格式化数据
                    for date, data_list in data_dict.items():
                        formatted_list = []
                        for item in data_list:
                            formatted_item = self.utils.validate_and_format_data(item)
                            if formatted_item:
                                formatted_list.append(formatted_item)
                                
                        if formatted_list:
                            if date in all_data:
                                all_data[date].extend(formatted_list)
                            else:
                                all_data[date] = formatted_list
                            processed_count += 1
                            
                    # 处理完成后，将文件移动到归档目录
                    if self.utils.move_excel(file_path, "池州华宇"):
                        self.logger.info(f"文件已归档: {file}")
                    
                except Exception as e:
                    self.logger.error(f"处理文件失败: {file} - {str(e)}")
                    continue
                    
            # 处理完成后的统计
            if excel_count == 0:
                self.logger.warning(f"目录中没有找到Excel文件: {attachment_folder}")
                return None
                
            if processed_count == 0:
                self.logger.warning("没有成功处理任何文件")
                return None
                
            self.logger.info(f"处理完成 - 总文件数: {excel_count}, 成功处理: {processed_count}")
            delivery_date = list(all_data.keys())[0]
            self.utils.save_json(all_data,f"池州华宇_{delivery_date}.json" , "池州华宇")
            return all_data
            
        except Exception as e:
            self.logger.error(f"处理池州华宇送货单失败: {str(e)}")
            return None
            
    def _process_huayu_return_dict(self, excel_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """
        处理池州华宇的送货单Excel文件并返回数据字典
        
        处理说明:
        1. 固定读取'Page 1'工作表
        2. 日期位置固定在P4单元格
        3. 从第7行开始读取数据，直到遇到'TOTAL'行
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            Dict[str, List[Dict[str, Any]]]: 按日期组织的数据字典
        """
        try:
            # 加载Excel工作簿，data_only=True表示读取值而不是公式
            wb = load_workbook(excel_path, data_only=True)
            sheet = wb['Page 1']  # 固定使用Page 1工作表
            
            data_dict = {}
            # 从固定位置(P4)获取日期
            date_str = sheet['P4'].value
            delivery_date = self.utils.format_date(str(date_str))
            
            if not delivery_date:
                self.logger.error("无法获取送货日期，跳过处理")
                return {}
                
            data_list = []
            
            # 从第8行开始遍历数据，直到遇到TOTAL行
            for row in range(8, sheet.max_row + 1):
                # 检查是否到达表格末尾（TOTAL行）
                if sheet[f'N{row}'].value == 'TOTAL':
                    break
                    
                # 跳过空行（以订单号是否存在为判断依据）
                if not sheet[f'D{row}'].value:
                    continue
                    
                try:
                    # 提取每行数据并构建数据字典
                    row_data = {
                        "送货日期": delivery_date,
                        "订单号": str(sheet[f'D{row}'].value or ''),
                        "品名": str(sheet[f'F{row}'].value or ''),
                        "封装形式": str(sheet[f'K{row}'].value or ''),
                        "打印批号": str(sheet[f'J{row}'].value or ''),
                        "数量": int(sheet[f'Q{row}'].value or 0),
                        "晶圆名称": str(sheet[f'O{row}'].value or ''),
                        "晶圆批号": str(sheet[f'L{row}'].value or ''),
                        "供应商": "池州华宇"
                    }
                    
                    # 验证和格式化数据
                    formatted_data = self.utils.validate_and_format_data(row_data)
                    if formatted_data:
                        data_list.append(formatted_data)
                    
                except Exception as e:
                    self.logger.error(f"处理第 {row} 行数据时出错: {str(e)}")
                    continue
                    
            # 如果有数据，则添加到返回字典中
            if data_list:            
                data_dict[delivery_date] = data_list              
            return data_dict
            
        except Exception as e:
            self.logger.error(f"提取送货单数据失败: {Path(excel_path).name} - {str(e)}")
            return {} 