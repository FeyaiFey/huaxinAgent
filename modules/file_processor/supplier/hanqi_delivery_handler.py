"""
汉旗供应商Excel处理器
处理汉旗供应商的送货单Excel文件
"""

import os
import xlrd
from typing import Dict, List, Optional, Any
from pathlib import Path
from openpyxl import load_workbook

from .base_delivery_handler import BaseDeliveryExcelHandler
from utils.logger import Logger

class HanQiHandler(BaseDeliveryExcelHandler):
    """
    汉旗供应商Excel处理器

    处理说明:
    1. 固定读取'Page 1'工作表
    2. 日期位置固定在P4单元格
    3. 从第7行开始读取数据，直到遇到'TOTAL'行
    4. 跳过空行（以订单号是否存在为判断依据）
    5. 验证和格式化数据
    6. 将数据保存到JSON文件
    7. 将Excel文件移动到归档目录
    """
    def __init__(self):
        self.logger = Logger(__name__)
    

    def process(self, match_result: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """
        处理汉旗的Excel文件
        
        Args:
            match_result: 规则引擎匹配结果
            
        Returns:
            Dict: 处理结果
        """
        try:
            # 用于存储所有数据
            all_data: Dict[str, List[Dict[str, Any]]] = {}

            # 确保目录存在
            attachment_folder = match_result.get('actions').get('attachment_folder')
            if not attachment_folder:
                self.logger.error(f"邮件数据中缺少附件目录")
                return None
                
            if not os.path.exists(attachment_folder):
                try:
                    os.makedirs(attachment_folder)
                    self.logger.debug(f"已创建目录: {attachment_folder}")
                except Exception as e:
                    self.logger.error(f"创建目录失败: {attachment_folder} - {str(e)}")
                    return None
                
            # 遍历目录下的所有文件
            excel_count = 0
            processed_count = 0

            for file in os.listdir(attachment_folder):
                if not file.lower().endswith(('.xlsx', '.xls')):
                    continue
                    
                file_path = os.path.join(attachment_folder, file)
                try:
                    self.logger.info(f"开始处理汉旗送货单: {Path(file_path).name}")
                    data_dict = self._process_hanqi_return_dict(file_path)
                    excel_count += 1
                    
                    if not data_dict:
                        self.logger.warning(f"文件处理未返回数据: {file}")
                        continue
                        
                    # 验证和格式化数据
                    for date, data_list in data_dict.items():
                        formatted_list = []
                        for item in data_list:
                            formatted_item = self.utils.validate_and_format_data(item)
                            if formatted_item:
                                formatted_list.append(formatted_item)
                                
                        if formatted_list:
                            if date in all_data:
                                all_data[date].extend(formatted_list)
                            else:
                                all_data[date] = formatted_list
                            processed_count += 1
                            
                    # 处理完成后，将文件移动到归档目录
                    if self.utils.move_excel(file_path, "山东汉旗"):
                        self.logger.debug(f"文件已归档: {file}")
                    
                except Exception as e:
                    self.logger.error(f"处理文件失败: {file} - {str(e)}")
                    continue
                    
            # 处理完成后的统计
            if excel_count == 0:
                self.logger.warning(f"目录中没有找到Excel文件: {attachment_folder}")
                return None
                
            if processed_count == 0:
                self.logger.warning("没有成功处理任何文件")
                return None
                
            self.logger.info(f"处理完成 - 总文件数: {excel_count}, 成功处理: {processed_count}")
            delivery_date = list(all_data.keys())[0]
            self.utils.save_json(all_data,f"山东汉旗_{delivery_date}.json" , "山东汉旗")
            return all_data
            
        except Exception as e:
            self.logger.error(f"处理汉旗送货单失败: {str(e)}")
            return None
            
    def _process_hanqi_return_dict(self, excel_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """
        处理山东汉旗的送货单Excel文件并返回数据字典
        
        处理说明:
        1. 遍历所有工作表
        2. 每个工作表的日期位置在G3单元格，格式为"日期:YYYY-MM-DD"
        3. 从第6行开始读取数据，直到遇到'Total'行
        4. 只处理日期大于上次处理日期的数据
        5. 支持旧版 .xls 和新版 .xlsx 格式
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            Dict[str, List[Dict[str, Any]]]: 按日期组织的数据字典
        """
        try:
            # 获取最后处理日期
            last_process_date = self.utils.get_last_process_date("山东汉旗")
            self.logger.info(f"山东汉旗最后处理日期: {last_process_date}")
            
            data_dict = {}
            max_processed_date = last_process_date  # 用于记录本次处理的最大日期

            if self.utils.is_xls_file(excel_path):
                # 使用 xlrd 处理 .xls 文件
                workbook = xlrd.open_workbook(excel_path)
                
                # 遍历所有工作表
                for sheet in workbook.sheets():
                    # 获取日期单元格内容 (G3 对应的是 row=2, col=6)
                    date_cell = sheet.cell_value(2, 6)
                    
                    # 检查日期单元格格式是否正确
                    if not date_cell or '日期：' not in str(date_cell):
                        continue
                        
                    # 提取并转换日期
                    delivery_date = self.utils.format_date(str(date_cell).split('日期：')[-1].strip())
                    if not delivery_date:
                        continue
                        
                    # 检查日期是否大于最后处理日期
                    if self.utils.compare_dates(delivery_date, last_process_date) <= 0:
                        self.logger.debug(f"跳过已处理的日期: {delivery_date}")
                        continue
                        
                    # 更新最大处理日期
                    if self.utils.compare_dates(delivery_date, max_processed_date) > 0:
                        max_processed_date = delivery_date
                        
                    data_list = []
                    
                    # 从第7行开始读取数据 (索引从0开始，所以是6)
                    for row in range(6, sheet.nrows):
                        # 检查是否到达表格末尾（Total行）
                        if 'Total' in str(sheet.cell_value(row, 7)):  # H列对应索引7
                            break
                        # 跳过空行
                        if not sheet.cell_value(row, 4):  # E列对应索引4
                            continue
                            
                        try:
                            # 提取每行数据
                            row_data = {
                                "送货日期": delivery_date,
                                "订单号": str(sheet.cell_value(row, 4)),  # E列
                                "品名": str(sheet.cell_value(row, 2)),    # C列
                                "封装形式": str(sheet.cell_value(row, 7)), # H列
                                "打印批号": str(sheet.cell_value(row, 5)), # F列
                                "数量": int(float(sheet.cell_value(row, 8)) or 0),  # I列
                                "晶圆名称": str(sheet.cell_value(row, 1)), # B列
                                "晶圆批号": str(sheet.cell_value(row, 3)), # D列
                                "供应商": "山东汉旗"
                            }
                            data_list.append(row_data)
                        except Exception as e:
                            self.logger.error(f"处理第 {row + 1} 行数据时出错: {str(e)}")
                            continue
                            
                    # 合并相同日期的数据
                    if data_list:
                        if delivery_date in data_dict:
                            data_dict[delivery_date].extend(data_list)
                        else:
                            data_dict[delivery_date] = data_list
            else:
                # 使用 openpyxl 处理 .xlsx 文件
                wb = load_workbook(excel_path, data_only=True)
                
                # 遍历所有工作表
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    # 获取日期单元格内容
                    date_cell = sheet['G3'].value
                    
                    # 检查日期单元格格式是否正确
                    if not date_cell or '日期:' not in str(date_cell):
                        continue
                        
                    # 提取并转换日期
                    delivery_date = self.utils.format_date(str(date_cell).split('日期:')[-1].strip())
                    if not delivery_date:
                        continue
                        
                    # 检查日期是否大于最后处理日期
                    if self.utils.compare_dates(delivery_date, last_process_date) <= 0:
                        self.logger.debug(f"跳过已处理的日期: {delivery_date}")
                        continue
                        
                    # 更新最大处理日期
                    if self.utils.compare_dates(delivery_date, max_processed_date) > 0:
                        max_processed_date = delivery_date
                        
                    data_list = []
                    
                    # 从第6行开始读取数据
                    for row in range(6, sheet.max_row + 1):
                        # 检查是否到达表格末尾（Total行）
                        if sheet[f'H{row}'].value == 'Total':
                            break
                            
                        # 跳过空行
                        if not sheet[f'E{row}'].value:
                            continue
                            
                        try:
                            # 提取每行数据
                            row_data = {
                                "送货日期": delivery_date,
                                "订单号": sheet[f'E{row}'].value,
                                "品名": sheet[f'C{row}'].value,
                                "封装形式": sheet[f'H{row}'].value,
                                "打印批号": sheet[f'F{row}'].value,
                                "数量": int(sheet[f'I{row}'].value or 0),
                                "晶圆名称": sheet[f'B{row}'].value,
                                "晶圆批号": sheet[f'D{row}'].value,
                                "供应商": "山东汉旗"
                            }
                            data_list.append(row_data)
                        except Exception as e:
                            self.logger.error(f"处理第 {row} 行数据时出错: {str(e)}")
                            continue
                            
                    # 合并相同日期的数据
                    if data_list:
                        if delivery_date in data_dict:
                            data_dict[delivery_date].extend(data_list)
                        else:
                            data_dict[delivery_date] = data_list
                            
            # 所有sheet处理完成后，更新最后处理日期
            if data_dict and self.utils.compare_dates(max_processed_date, last_process_date) > 0:
                self.utils.update_last_process_date("山东汉旗", max_processed_date)
                self.logger.debug(f"更新山东汉旗最后处理日期为: {max_processed_date}")
                
            return data_dict
            
        except Exception as e:
            self.logger.error(f"处理山东汉旗送货单失败: {str(e)}")
            return {}