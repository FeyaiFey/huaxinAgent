"""
供应商Excel处理器工具类
提供通用的文件处理、日期处理和数据验证功能
"""

import os
import json
import shutil
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from typing import Dict, List, Optional, Any
from datetime import datetime

from utils.logger import Logger
from utils.helpers import load_yaml

class SupplierUtils:
    """供应商Excel处理器工具类"""
    
    def __init__(self):
        """
        初始化工具类
        """
        self.logger = Logger(__name__)
        self.delivery_config = load_yaml('config/delivery_json_format.yaml')
        self.wip_fields_config = load_yaml('config/wip_fields.yaml')
        self.settings = load_yaml('config/settings.yaml')
        
    def save_json(self, data: List[Dict[str, Any]], filename: str, supplier: str) -> Optional[str]:
        """
        保存JSON数据到指定位置
        
        Args:
            data: 要保存的数据列表
            filename: 文件名
            supplier: 供应商标识
            
        Returns:
            Optional[str]: 保存成功返回文件路径，失败返回None
        """
        try:
            # 从配置文件获取基础路径
            base_dir = self.settings['file_management']['delivery_json_save_dir']
            
            # 确保输出目录存在
            output_dir = os.path.join(base_dir, supplier)
            os.makedirs(output_dir, exist_ok=True)
            
            # 构建完整的文件路径
            json_path = os.path.join(output_dir, filename)
            
            # 保存JSON文件
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
                
            self.logger.info(f"已保存JSON数据 [{json_path}]")
            return json_path
        except Exception as e:
            self.logger.error(f"保存JSON数据失败 [{filename}]: {str(e)}")
            return None
            
    def move_excel(self, excel_path: str, supplier: str) -> bool:
        """
        移动Excel文件到指定位置
        
        Args:
            excel_path: Excel文件路径
            supplier: 供应商标识
            
        Returns:
            bool: 移动成功返回True，失败返回False
        """
        try:
            if not excel_path or not os.path.exists(excel_path):
                self.logger.error(f"源Excel文件不存在: {excel_path}")
                return False

            # 确保归档目录存在
            archive_dir = os.path.join('attachments/delivery_notes', supplier)
            os.makedirs(archive_dir, exist_ok=True)
            
            # 构建目标路径
            filename = os.path.basename(excel_path)
            target_path = os.path.join(archive_dir, filename)
            
            # 移动文件
            shutil.move(excel_path, target_path)
            self.logger.debug(f"已归档Excel文件 [{target_path}]")
            return True
        except Exception as e:
            self.logger.error(f"归档Excel文件失败 [{excel_path}]: {str(e)}")
            return False
            
    def format_date(self, date_str: str, from_format: bool = True) -> Optional[str]:
        """
        日期格式转换
        
        Args:
            date_str: 日期字符串
            from_format: True表示转换为YYYY-MM-DD格式，False表示转换为YYYYMMDD格式
            
        Returns:
            Optional[str]: 转换后的日期字符串，如果转换失败则返回None
        """
        try:
            # 处理特殊情况
            if date_str in ["0000-00-00", "00000000"]:
                return "0000-00-00" if from_format else "00000000"
                
            if from_format:
                # 将其他格式转换为YYYY-MM-DD
                date_formats = [
                    '%Y%m%d',           # YYYYMMDD
                    '%Y-%m-%d',         # YYYY-MM-DD
                    '%Y/%m/%d',         # YYYY/MM/DD
                    '%Y.%m.%d',         # YYYY.MM.DD
                    '%Y年%m月%d日',      # YYYY年MM月DD日
                    '%Y-%m-%d %H:%M:%S', # YYYY-MM-DD HH:MM:SS
                    '%Y/%m/%d %H:%M:%S', # YYYY/MM/DD HH:MM:SS
                    '%Y-%m-%d %H:%M',    # YYYY-MM-DD HH:MM
                    '%Y/%m/%d %H:%M'     # YYYY/MM/DD HH:MM
                ]
                
                for fmt in date_formats:
                    try:
                        date_obj = datetime.strptime(str(date_str).strip(), fmt)
                        return date_obj.strftime('%Y-%m-%d')
                    except ValueError:
                        continue
            else:
                # 将YYYY-MM-DD转换为YYYYMMDD（用于比较和存储）
                try:
                    # 如果包含时间，先提取日期部分
                    if ' ' in str(date_str):
                        date_str = str(date_str).split(' ')[0]
                        
                    if '-' in date_str:
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                        return date_obj.strftime('%Y%m%d')
                    else:
                        date_obj = datetime.strptime(date_str, '%Y%m%d')
                        return date_str
                except ValueError:
                    pass
                    
            self.logger.warning(f"无法解析日期格式: {date_str}")
            return None
            
        except Exception as e:
            self.logger.error(f"处理日期时出错: {str(e)}")
            return None
            
    def compare_dates(self, date1: str, date2: str) -> int:
        """
        比较两个日期的大小
        
        Args:
            date1: 第一个日期（YYYY-MM-DD格式）
            date2: 第二个日期（YYYY-MM-DD格式）
            
        Returns:
            int: 如果date1 > date2返回1，如果date1 < date2返回-1，如果相等返回0
        """
        try:
            # 处理特殊情况
            if date1 == date2:
                return 0
                
            if date1 == "0000-00-00":
                return -1
                
            if date2 == "0000-00-00":
                return 1
                
            # 转换为YYYYMMDD格式进行比较
            date1_fmt = self.format_date(date1, False)
            date2_fmt = self.format_date(date2, False)
            
            if not date1_fmt or not date2_fmt:
                return 0
                
            if date1_fmt > date2_fmt:
                return 1
            elif date1_fmt < date2_fmt:
                return -1
            else:
                return 0
                
        except Exception as e:
            self.logger.error(f"比较日期时出错: {str(e)}")
            return 0
            
    def validate_and_format_data(self, data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """
        验证和格式化数据，确保符合配置文件中定义的格式
        
        Args:
            data: 原始数据字典
            
        Returns:
            Optional[Dict[str, Any]]: 格式化后的数据字典，如果验证失败则返回None
        """
        try:
            # 获取字段定义
            fields = self.delivery_config['fields']
            formatted_data = {}
            
            # 验证每个字段
            for field in fields:
                field_name = field['name']
                field_type = field['type']
                required = field['required']
                
                # 获取字段值
                value = data.get(field_name)
                
                # 检查必填字段
                if required and value is None:
                    self.logger.error(f"缺少必填字段: {field_name}")
                    return None
                    
                # 如果字段不存在且非必填，使用默认值
                if value is None:
                    formatted_data[field_name] = ""
                    continue
                    
                # 根据字段类型进行格式化
                try:
                    if field_type == "date":
                        # 确保日期格式为YYYY-MM-DD
                        formatted_data[field_name] = self.format_date(str(value))
                    elif field_type == "integer":
                        # 确保数字字段为整数
                        formatted_data[field_name] = int(float(value)) if value else 0
                    elif field_type == "string":
                        # 确保字符串字段为字符串类型，并去除首尾空白
                        formatted_data[field_name] = str(value).strip() if value else ""
                    else:
                        formatted_data[field_name] = value
                except Exception as e:
                    self.logger.error(f"字段 {field_name} 格式化失败: {str(e)}")
                    return None
                    
            return formatted_data
            
        except Exception as e:
            self.logger.error(f"数据验证和格式化失败: {str(e)}")
            return None
            
    def get_last_process_date(self, supplier: str) -> str:
        """
        获取供应商最后一次处理的送货日期
        
        Args:
            supplier: 供应商标识
            
        Returns:
            str: 最后处理的日期（YYYY-MM-DD格式），如果没有记录则返回'0000-00-00'
        """
        try:
            # 读取记录文件
            process_dates_file = os.path.join("config", "process_dates.json")
            if not os.path.exists(process_dates_file):
                # 如果文件不存在，创建默认内容
                default_content = {
                    "山东汉旗": "0000-00-00",
                    "池州华宇": "0000-00-00",
                    "江苏芯丰": "0000-00-00"
                }
                with open(process_dates_file, 'w', encoding='utf-8') as f:
                    json.dump(default_content, f, ensure_ascii=False, indent=2)
                return "0000-00-00"
                
            # 读取日期记录
            with open(process_dates_file, 'r', encoding='utf-8') as f:
                dates = json.load(f)
                return dates.get(supplier, "0000-00-00")
                
        except Exception as e:
            self.logger.error(f"获取最后处理日期失败: {str(e)}")
            return "0000-00-00"
            
    def update_last_process_date(self, supplier: str, date: str) -> bool:
        """
        更新供应商最后一次处理的送货日期
        
        Args:
            supplier: 供应商标识
            date: 新的处理日期（YYYY-MM-DD格式）
            
        Returns:
            bool: 更新成功返回True，失败返回False
        """
        try:
            process_dates_file = os.path.join("config", "process_dates.json")
            
            # 读取现有记录
            with open(process_dates_file, 'r', encoding='utf-8') as f:
                dates = json.load(f)
                
            # 更新日期
            current_date = dates.get(supplier, "0000-00-00")
            # 转换为YYYYMMDD格式进行比较
            if self.format_date(date, False) > self.format_date(current_date, False):
                dates[supplier] = date
                
                # 保存更新后的记录
                with open(process_dates_file, 'w', encoding='utf-8') as f:
                    json.dump(dates, f, ensure_ascii=False, indent=2)
                    
                self.logger.debug(f"已更新{supplier}的最后处理日期: {date}")
                return True
                
            return False
            
        except Exception as e:
            self.logger.error(f"更新最后处理日期失败: {str(e)}")
            return False
            
    def is_xls_file(self, file_path: str) -> bool:
        """
        检查文件是否为旧版 .xls 格式
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 如果是 .xls 文件返回 True，否则返回 False
        """
        return file_path.lower().endswith('.xls') 
    
    def copy_to_gzjc(self, supplier: str) -> bool:
        """
        将JSON数据写入到工作进程Excel表尾
        
        Args:
            supplier: 供应商名称
            
        Returns:
            bool: 写入成功返回True，失败返回False
        """
        try:
            # 检查工作进程Excel是否可写
            gzjc_path = self.settings['file_management']['gzjc_path']
            if not gzjc_path:
                self.logger.error("工作进程Excel文件路径未配置")
                return False
                
            # 检查文件是否存在
            if not os.path.exists(gzjc_path):
                self.logger.error(f"工作进程Excel文件不存在: {gzjc_path}")
                return False
                
            # 检查文件是否可写
            try:
                with open(gzjc_path, 'r+b') as f:
                    pass
            except PermissionError:
                self.logger.error(f"工作进程Excel文件无法写入: {gzjc_path}")
                return False
            except Exception as e:
                self.logger.error(f"工作进程Excel文件访问失败: {str(e)}")
                return False
            
            # 加载工作簿
            wb = load_workbook(gzjc_path)
            if '入库记录' not in wb.sheetnames:
                self.logger.error("工作进程Excel中未找到'入库记录'表")
                return False
                
            ws = wb['入库记录']
            last_row = ws.max_row + 1
                
            # 获取JSON文件目录
            json_dir = os.path.join(self.settings['file_management']['delivery_json_save_dir'], supplier)
            if not os.path.exists(json_dir):
                self.logger.error(f"JSON文件目录不存在: {json_dir}")
                return False
                
            # 遍历JSON文件
            success = True
            for json_file in os.listdir(json_dir):
                if not json_file.endswith('.json') or 'success_to_gzjc' in json_file:
                    continue
                    
                json_path = os.path.join(json_dir, json_file)
                try:
                    # 读取 JSON 数据
                    with open(json_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        
                    # 处理每个日期下的数据
                    for date, records in data.items():
                        if not isinstance(records, list):
                            continue
                            
                        # 将记录转换为 DataFrame
                        df = pd.DataFrame(records)
                        
                        # 写入数据
                        for _, row in df.iterrows():
                            # 转换为日期对象
                            date_obj = datetime.strptime(row['送货日期'], '%Y-%m-%d')
                            
                            # 写入数据并设置格式
                            cells = [
                                (1, date_obj), (2, row['订单号']), (3, row['品名']),
                                (4, row['晶圆名称']), (5, row['晶圆批号']), (6, row['封装形式']),
                                (7, row['数量']), (8, row['打印批号']), (9, row['供应商']),
                                (10, "合金丝")
                            ]
                            
                            for col, value in cells:
                                cell = ws.cell(row=last_row, column=col, value=value)
                                # 设置字体
                                cell.font = openpyxl.styles.Font(name='Times New Roman')
                                # 设置对齐方式
                                if col in [6, 7]:  # F和G列居中对齐
                                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                                else:  # 其他列左对齐
                                    cell.alignment = openpyxl.styles.Alignment(horizontal='left')
                                
                                # 设置日期格式
                                if col == 1:  # 日期列
                                    cell.number_format = 'yyyy/m/d'
                            
                            last_row += 1
                    
                    # 保存工作簿
                    wb.save(gzjc_path)
                    
                    # 重命名 JSON 文件
                    new_name = os.path.splitext(json_file)[0] + "_success_to_gzjc.json"
                    os.rename(json_path, os.path.join(json_dir, new_name))
                    self.logger.debug(f"已处理并重命名文件: {json_file}")
                    
                except Exception as e:
                    self.logger.error(f"处理文件 {json_file} 时出错: {str(e)}")
                    success = False
                    continue
            
            return success
            
        except Exception as e:
            self.logger.error(f"写入工作进程Excel失败: {str(e)}")
            return False